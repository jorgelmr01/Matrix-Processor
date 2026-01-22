#!/usr/bin/env python3
"""
Procesador de Matrices - Aplicación web para crear matrices de intersección desde archivos Excel/CSV.
Ejecutar con: python app.py
"""

import os
import sys
import json
import webbrowser
import threading
from datetime import datetime
from io import BytesIO
from http.server import HTTPServer, SimpleHTTPRequestHandler
import urllib.parse

# Check for required packages and install if missing
def check_dependencies():
    required = ['openpyxl', 'pandas']
    missing = []
    for pkg in required:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pkg)
    
    if missing:
        print(f"Instalando paquetes requeridos: {', '.join(missing)}")
        import subprocess
        subprocess.check_call([sys.executable, '-m', 'pip', 'install'] + missing, 
                            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        print("¡Paquetes instalados exitosamente!")

check_dependencies()

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Store uploaded files and processing state
app_state = {
    'files': {},
    'file_data': [],
    'filter_file': None
}

class MatrixProcessorHandler(SimpleHTTPRequestHandler):
    # Increase timeout for large file uploads
    timeout = 300  # 5 minutes timeout
    
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=os.path.dirname(os.path.abspath(__file__)), **kwargs)
    
    def log_message(self, format, *args):
        pass  # Suppress logging
    
    def send_json(self, data, status=200):
        response = json.dumps(data).encode('utf-8')
        self.send_response(status)
        self.send_header('Content-Type', 'application/json')
        self.send_header('Content-Length', len(response))
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(response)
    
    def send_file_download(self, data, filename):
        self.send_response(200)
        self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
        self.send_header('Content-Length', len(data))
        self.send_header('Access-Control-Allow-Origin', '*')
        self.end_headers()
        self.wfile.write(data)
    
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def end_headers(self):
        # Add no-cache headers to prevent browser caching issues
        self.send_header('Cache-Control', 'no-cache, no-store, must-revalidate')
        self.send_header('Pragma', 'no-cache')
        self.send_header('Expires', '0')
        super().end_headers()
    
    def do_GET(self):
        if self.path == '/':
            self.path = '/index.html'
        
        if self.path == '/api/status':
            self.send_json({'status': 'ok', 'files': len(app_state['file_data'])})
        else:
            super().do_GET()
    
    def do_POST(self):
        content_length = int(self.headers.get('Content-Length', 0))
        
        if self.path == '/api/upload':
            self.handle_upload(content_length)
        elif self.path == '/api/upload-filter':
            self.handle_upload_filter(content_length)
        elif self.path == '/api/process':
            self.handle_process(content_length)
        elif self.path == '/api/compute':
            self.handle_compute(content_length)
        elif self.path == '/api/export':
            self.handle_export(content_length)
        elif self.path == '/api/reset':
            app_state['files'] = {}
            app_state['file_data'] = []
            app_state['filter_file'] = None
            self.send_json({'status': 'ok'})
        else:
            self.send_json({'error': 'No encontrado'}, 404)
    
    def handle_upload(self, content_length):
        """Handle file upload"""
        content_type = self.headers.get('Content-Type', '')
        
        if 'multipart/form-data' in content_type:
            # Parse multipart form data
            boundary = content_type.split('boundary=')[1].encode()
            
            # Read body in chunks to handle large uploads
            body = b''
            remaining = content_length
            chunk_size = 1024 * 1024  # 1MB chunks
            while remaining > 0:
                to_read = min(chunk_size, remaining)
                chunk = self.rfile.read(to_read)
                if not chunk:
                    break
                body += chunk
                remaining -= len(chunk)
            
            parts = body.split(b'--' + boundary)
            files_processed = []
            
            for part in parts:
                if b'filename="' in part:
                    # Extract filename
                    header_end = part.find(b'\r\n\r\n')
                    header = part[:header_end].decode('utf-8', errors='ignore')
                    filename_start = header.find('filename="') + 10
                    filename_end = header.find('"', filename_start)
                    filename = header[filename_start:filename_end]
                    
                    # Extract file content
                    file_content = part[header_end + 4:]
                    if file_content.endswith(b'\r\n'):
                        file_content = file_content[:-2]
                    
                    if filename and file_content:
                        app_state['files'][filename] = file_content
                        files_processed.append(filename)
            
            # Process all uploaded files
            app_state['file_data'] = []
            errors = []
            for filename, content in app_state['files'].items():
                try:
                    file_info = self.process_file(filename, content)
                    app_state['file_data'].append(file_info)
                except Exception as e:
                    errors.append(f'{filename}: {str(e)}')
                    print(f"Error processing {filename}: {e}")
            
            if errors and not app_state['file_data']:
                # All files failed
                self.send_json({'error': f'Error al procesar archivos:\n' + '\n'.join(errors)}, 400)
                return
            elif errors:
                # Some files failed but some succeeded
                print(f"Warnings: {len(errors)} files had errors")
            
            self.send_json({'status': 'ok', 'files': app_state['file_data']})
        else:
            self.send_json({'error': 'Tipo de contenido inválido'}, 400)
    
    def process_file(self, filename, content):
        """Process an Excel or CSV file"""
        file_info = {
            'fileName': filename,
            'fileType': 'csv' if filename.endswith('.csv') else 'excel',
            'sheets': []
        }
        
        try:
            if filename.endswith('.csv'):
                df = pd.read_csv(BytesIO(content))
                # Trim headers and data
                df.columns = [str(col).strip() for col in df.columns]
                headers = df.columns.tolist()
                data = df.fillna('').astype(str).apply(lambda x: x.str.strip()).to_dict('records')
                file_info['sheets'].append({
                    'name': 'Sheet1',
                    'headers': headers,
                    'data': data
                })
            else:
                xlsx = pd.ExcelFile(BytesIO(content))
                for sheet_name in xlsx.sheet_names:
                    df = pd.read_excel(xlsx, sheet_name=sheet_name)
                    # Trim headers and data
                    df.columns = [str(col).strip() for col in df.columns]
                    headers = df.columns.tolist()
                    data = df.fillna('').astype(str).apply(lambda x: x.str.strip()).to_dict('records')
                    file_info['sheets'].append({
                        'name': sheet_name,
                        'headers': headers,
                        'data': data
                    })
        except Exception as e:
            raise Exception(f'Error al leer archivo: {str(e)}')
        
        return file_info
    
    def handle_upload_filter(self, content_length):
        """Handle filter file upload"""
        content_type = self.headers.get('Content-Type', '')
        
        if 'multipart/form-data' in content_type:
            boundary = content_type.split('boundary=')[1].encode()
            
            # Read body in chunks to handle large uploads
            body = b''
            remaining = content_length
            chunk_size = 1024 * 1024  # 1MB chunks
            while remaining > 0:
                to_read = min(chunk_size, remaining)
                chunk = self.rfile.read(to_read)
                if not chunk:
                    break
                body += chunk
                remaining -= len(chunk)
            
            parts = body.split(b'--' + boundary)
            
            for part in parts:
                if b'filename="' in part:
                    header_end = part.find(b'\r\n\r\n')
                    header = part[:header_end].decode('utf-8', errors='ignore')
                    filename_start = header.find('filename="') + 10
                    filename_end = header.find('"', filename_start)
                    filename = header[filename_start:filename_end]
                    
                    file_content = part[header_end + 4:]
                    if file_content.endswith(b'\r\n'):
                        file_content = file_content[:-2]
                    
                    if filename and file_content:
                        try:
                            file_info = self.process_file(filename, file_content)
                            app_state['filter_file'] = file_info
                            self.send_json({'status': 'ok', 'file': file_info})
                            return
                        except Exception as e:
                            self.send_json({'error': f'Error al procesar archivo de filtro: {str(e)}'}, 400)
                            return
            
            self.send_json({'error': 'No se encontró archivo'}, 400)
        else:
            self.send_json({'error': 'Tipo de contenido inválido'}, 400)
    
    def handle_process(self, content_length):
        """Return processed file data"""
        self.send_json({'files': app_state['file_data']})
    
    def handle_compute(self, content_length):
        """Compute matrices based on configuration"""
        body = self.rfile.read(content_length)
        config = json.loads(body.decode('utf-8'))
        
        # Extract filter config if present
        filter_config = config.get('filterConfig')
        filter_data = None
        source_mappings = {}
        filter_column_mappings = {}
        
        if filter_config and filter_config.get('enabled'):
            # Get all filter data (per column)
            all_filter_data = filter_config.get('allFilterData', {})
            if all_filter_data:
                # Convert to sets for efficient lookup
                filter_data = {}
                for col, data in all_filter_data.items():
                    filter_data[col] = set(v.lower().strip() for v in data.get('valuesLower', []))
            
            # Get source-specific column mappings (which column in data file to match)
            source_mappings = filter_config.get('sourceMappings', {})
            # Get filter column mappings (which column in filter file to use per source)
            filter_column_mappings = filter_config.get('filterColumnMappings', {})
        
        try:
            matrices = self.compute_matrices(
                config['fileData'],
                config['selectedTabs'],
                config['columnSelections'],
                config['matrixConfig'],
                filter_data,
                source_mappings,
                filter_column_mappings
            )
            self.send_json({'matrices': matrices})
        except Exception as e:
            self.send_json({'error': str(e)}, 400)
    
    def get_row_value(self, row, x_axis_columns):
        """Get combined row value from multiple columns"""
        parts = []
        for col in x_axis_columns:
            val = str(row.get(col, '')).strip()
            if val:
                parts.append(val)
        return ' | '.join(parts) if parts else ''
    
    def compute_matrices(self, file_data, selected_tabs, column_selections, matrix_config, filter_data=None, source_mappings=None, filter_column_mappings=None):
        """Compute intersection matrices with multi-column X axis support
        
        Args:
            filter_data: Dict mapping filter column name -> set of lowercase values
            source_mappings: Dict mapping source key to column name in data file for filter matching
            filter_column_mappings: Dict mapping source key to column name in filter file to use
        """
        if source_mappings is None:
            source_mappings = {}
        if filter_column_mappings is None:
            filter_column_mappings = {}
        
        matrices = []
        
        for config in matrix_config:
            if config.get('merge'):
                # Merge all sources into one matrix
                y_values = set()
                x_values = set()
                
                # Collect unique values
                for source in config['sources']:
                    file_idx = source['fileIndex']
                    sheet_name = source['sheetName']
                    key = f"{file_idx}-{sheet_name}"
                    
                    file = file_data[file_idx]
                    sheet = next((s for s in file['sheets'] if s['name'] == sheet_name), None)
                    if not sheet:
                        continue
                    
                    sel = column_selections.get(key, {})
                    y_col = sel.get('yAxis')
                    x_cols = sel.get('xAxisMultiple', [])
                    
                    # Fallback to single xAxis if xAxisMultiple not set
                    if not x_cols and sel.get('xAxis'):
                        x_cols = [sel.get('xAxis')]
                    
                    # Get filter column for this source (which column in data file to match)
                    source_filter_col = source_mappings.get(key)
                    # Get which filter file column to use for this source
                    filter_file_col = filter_column_mappings.get(key)
                    
                    # Get the filter values for this source's filter column
                    source_filter_values = None
                    if filter_data and filter_file_col and filter_file_col in filter_data:
                        source_filter_values = filter_data[filter_file_col]
                    
                    for row in sheet['data']:
                        y_val = str(row.get(y_col, '')).strip()
                        x_val = self.get_row_value(row, x_cols)
                        
                        if y_val:
                            y_values.add(y_val)
                        if x_val:
                            # Apply filter if present - using source-specific columns
                            if source_filter_values is None:
                                x_values.add(x_val)
                            elif source_filter_col:
                                # Use the mapped source column to get value and match against filter
                                filter_val = str(row.get(source_filter_col, '')).strip().lower()
                                if filter_val in source_filter_values:
                                    x_values.add(x_val)
                            else:
                                # Fallback: check if any filter value is in the row value
                                x_val_lower = x_val.lower()
                                if any(fv in x_val_lower for fv in source_filter_values):
                                    x_values.add(x_val)
                
                sorted_y = sorted(y_values)
                sorted_x = sorted(x_values)
                
                matrix_data = [[0] * len(sorted_y) for _ in range(len(sorted_x))]
                
                for source in config['sources']:
                    file_idx = source['fileIndex']
                    sheet_name = source['sheetName']
                    key = f"{file_idx}-{sheet_name}"
                    
                    file = file_data[file_idx]
                    sheet = next((s for s in file['sheets'] if s['name'] == sheet_name), None)
                    if not sheet:
                        continue
                    
                    sel = column_selections.get(key, {})
                    y_col = sel.get('yAxis')
                    x_cols = sel.get('xAxisMultiple', [])
                    
                    if not x_cols and sel.get('xAxis'):
                        x_cols = [sel.get('xAxis')]
                    
                    for row in sheet['data']:
                        y_val = str(row.get(y_col, '')).strip()
                        x_val = self.get_row_value(row, x_cols)
                        
                        if y_val and x_val:
                            if y_val in sorted_y and x_val in sorted_x:
                                y_idx = sorted_y.index(y_val)
                                x_idx = sorted_x.index(x_val)
                                matrix_data[x_idx][y_idx] = 1
                
                matrices.append({
                    'name': config['name'],
                    'yAxis': sorted_y,
                    'xAxis': sorted_x,
                    'data': matrix_data
                })
            else:
                # Create independent matrix for each source
                for source in config['sources']:
                    file_idx = source['fileIndex']
                    sheet_name = source['sheetName']
                    key = f"{file_idx}-{sheet_name}"
                    
                    file = file_data[file_idx]
                    sheet = next((s for s in file['sheets'] if s['name'] == sheet_name), None)
                    if not sheet:
                        continue
                    
                    sel = column_selections.get(key, {})
                    y_col = sel.get('yAxis')
                    x_cols = sel.get('xAxisMultiple', [])
                    
                    if not x_cols and sel.get('xAxis'):
                        x_cols = [sel.get('xAxis')]
                    
                    # Get filter column for this source (which column in data file to match)
                    source_filter_col = source_mappings.get(key)
                    # Get which filter file column to use for this source
                    filter_file_col = filter_column_mappings.get(key)
                    
                    # Get the filter values for this source's filter column
                    source_filter_values = None
                    if filter_data and filter_file_col and filter_file_col in filter_data:
                        source_filter_values = filter_data[filter_file_col]
                    
                    y_values = set()
                    x_values = set()
                    
                    for row in sheet['data']:
                        y_val = str(row.get(y_col, '')).strip()
                        x_val = self.get_row_value(row, x_cols)
                        
                        if y_val:
                            y_values.add(y_val)
                        if x_val:
                            # Apply filter if present - using source-specific columns
                            if source_filter_values is None:
                                x_values.add(x_val)
                            elif source_filter_col:
                                # Use the mapped source column to get value and match against filter
                                filter_val = str(row.get(source_filter_col, '')).strip().lower()
                                if filter_val in source_filter_values:
                                    x_values.add(x_val)
                            else:
                                # Fallback: check if any filter value is in the row value
                                x_val_lower = x_val.lower()
                                if any(fv in x_val_lower for fv in source_filter_values):
                                    x_values.add(x_val)
                    
                    sorted_y = sorted(y_values)
                    sorted_x = sorted(x_values)
                    
                    # matrix_data[x_idx][y_idx] - X is rows, Y is columns
                    matrix_data = [[0] * len(sorted_y) for _ in range(len(sorted_x))]
                    
                    for row in sheet['data']:
                        y_val = str(row.get(y_col, '')).strip()
                        x_val = self.get_row_value(row, x_cols)
                        
                        if y_val and x_val:
                            if y_val in sorted_y and x_val in sorted_x:
                                y_idx = sorted_y.index(y_val)
                                x_idx = sorted_x.index(x_val)
                                matrix_data[x_idx][y_idx] = 1
                    
                    matrices.append({
                        'name': config['name'],
                        'yAxis': sorted_y,
                        'xAxis': sorted_x,
                        'data': matrix_data
                    })
        
        return matrices
    
    def handle_export(self, content_length):
        """Export matrices to Excel with Consulta sheet"""
        body = self.rfile.read(content_length)
        data = json.loads(body.decode('utf-8'))
        matrices = data.get('matrices', [])
        
        try:
            wb = Workbook()
            wb.remove(wb.active)
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell_border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )
            one_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            
            # ========== CONSULTA SHEET ==========
            # Dynamic lookup with dropdown - one column per matrix
            lookup_ws = wb.create_sheet(title="Consulta")
            
            # Collect data per matrix: {matrix_name: {row_val: [perm1, perm2, ...]}}
            matrix_data_lookup = {}
            all_row_values = set()
            matrix_names = [m['name'][:30] for m in matrices]  # Truncate names for display
            
            for matrix in matrices:
                matrix_name = matrix['name'][:30]
                rows = matrix['xAxis']
                cols = matrix['yAxis']
                m_data = matrix['data']
                
                if matrix_name not in matrix_data_lookup:
                    matrix_data_lookup[matrix_name] = {}
                
                for row_idx, row_val in enumerate(rows):
                    all_row_values.add(row_val)
                    if row_idx < len(m_data):
                        perms = []
                        for col_idx, val in enumerate(m_data[row_idx]):
                            if val == 1 and col_idx < len(cols):
                                perms.append(cols[col_idx])
                        if perms:
                            if row_val not in matrix_data_lookup[matrix_name]:
                                matrix_data_lookup[matrix_name][row_val] = []
                            matrix_data_lookup[matrix_name][row_val].extend(perms)
            
            sorted_row_values = sorted(all_row_values, key=str.lower)
            
            # Calculate max permissions any user has in any matrix
            max_perms = 1
            for matrix_name in matrix_names:
                for row_val, perms in matrix_data_lookup.get(matrix_name, {}).items():
                    max_perms = max(max_perms, len(perms))
            
            num_result_rows = max(max_perms, 25)
            num_cols = len(matrix_names) + 1
            
            # Title
            lookup_ws.merge_cells(f'A1:{get_column_letter(num_cols)}1')
            title_cell = lookup_ws['A1']
            title_cell.value = "Consulta de Permisos por Usuario"
            title_cell.font = Font(bold=True, size=16, color="1E293B")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            lookup_ws.row_dimensions[1].height = 30
            
            # Subtitle
            lookup_ws.merge_cells(f'A2:{get_column_letter(num_cols)}2')
            subtitle_cell = lookup_ws['A2']
            subtitle_cell.value = "Selecciona un usuario del menú desplegable para ver sus permisos"
            subtitle_cell.font = Font(size=11, color="64748B")
            subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
            lookup_ws.row_dimensions[2].height = 25
            
            # Dropdown label
            lookup_ws['A4'] = "Seleccionar Usuario:"
            lookup_ws['A4'].font = Font(bold=True)
            lookup_ws['A4'].alignment = Alignment(horizontal="right", vertical="center")
            
            # Store dropdown values in a hidden column (column after all matrix columns + buffer)
            dropdown_col = num_cols + 3
            for idx, val in enumerate(sorted_row_values, start=1):
                lookup_ws.cell(row=idx, column=dropdown_col, value=val)
            lookup_ws.column_dimensions[get_column_letter(dropdown_col)].hidden = True
            
            # Create dropdown in B4
            if sorted_row_values:
                dv = DataValidation(
                    type="list",
                    formula1=f"${get_column_letter(dropdown_col)}$1:${get_column_letter(dropdown_col)}${len(sorted_row_values)}",
                    allow_blank=True
                )
                dv.error = "Por favor selecciona un valor de la lista"
                dv.errorTitle = "Valor inválido"
                dv.prompt = "Selecciona un usuario"
                dv.promptTitle = "Lista de usuarios"
                lookup_ws.add_data_validation(dv)
                dv.add(lookup_ws['B4'])
            
            lookup_ws['B4'].fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
            lookup_ws['B4'].border = Border(
                left=Side(style='medium', color='F59E0B'),
                right=Side(style='medium', color='F59E0B'),
                top=Side(style='medium', color='F59E0B'),
                bottom=Side(style='medium', color='F59E0B')
            )
            lookup_ws.column_dimensions['A'].width = 20
            lookup_ws.column_dimensions['B'].width = 45
            
            # Headers row - "#" column + one column per matrix
            lookup_ws['A6'] = "#"
            lookup_ws['A6'].font = header_font
            lookup_ws['A6'].fill = header_fill
            lookup_ws['A6'].alignment = header_align
            lookup_ws['A6'].border = cell_border
            
            for col_idx, matrix_name in enumerate(matrix_names, start=2):
                cell = lookup_ws.cell(row=6, column=col_idx, value=matrix_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = cell_border
                lookup_ws.column_dimensions[get_column_letter(col_idx)].width = 35
            
            # Build lookup data for each matrix in hidden columns
            # Structure: For each matrix, create columns with Key (user|index) and Value (permission)
            lookup_start_col = dropdown_col + 2
            matrix_lookup_ranges = {}
            
            for matrix_idx, matrix_name in enumerate(matrix_names):
                key_col = lookup_start_col + (matrix_idx * 2)
                val_col = key_col + 1
                
                row_num = 1
                matrix_perms = matrix_data_lookup.get(matrix_name, {})
                
                for row_val in sorted_row_values:
                    if row_val in matrix_perms:
                        for perm_idx, perm in enumerate(sorted(set(matrix_perms[row_val])), start=1):
                            # Key format: "user_value|index"
                            lookup_ws.cell(row=row_num, column=key_col, value=f"{row_val}|{perm_idx}")
                            lookup_ws.cell(row=row_num, column=val_col, value=perm)
                            row_num += 1
                
                # Store range info
                if row_num > 1:
                    matrix_lookup_ranges[matrix_name] = {
                        'key_col': get_column_letter(key_col),
                        'val_col': get_column_letter(val_col),
                        'last_row': row_num - 1
                    }
                
                # Hide these columns
                lookup_ws.column_dimensions[get_column_letter(key_col)].hidden = True
                lookup_ws.column_dimensions[get_column_letter(val_col)].hidden = True
            
            # Create result rows with VLOOKUP formulas
            for i in range(num_result_rows):
                result_row = 7 + i
                
                # Row number
                cell = lookup_ws.cell(row=result_row, column=1, value=i + 1)
                cell.border = cell_border
                cell.alignment = Alignment(horizontal="center")
                cell.font = Font(color="94A3B8")
                
                # For each matrix, create VLOOKUP formula
                for matrix_idx, matrix_name in enumerate(matrix_names, start=2):
                    if matrix_name in matrix_lookup_ranges:
                        info = matrix_lookup_ranges[matrix_name]
                        key_col = info['key_col']
                        val_col = info['val_col']
                        last_row = info['last_row']
                        
                        # VLOOKUP formula: =IFERROR(VLOOKUP($B$4&"|"&row_num, key_col:val_col, 2, FALSE), "")
                        formula = f'=IFERROR(VLOOKUP($B$4&"|"&{i+1},${key_col}$1:${val_col}${last_row},2,FALSE),"")'
                        cell = lookup_ws.cell(row=result_row, column=matrix_idx, value=formula)
                    else:
                        cell = lookup_ws.cell(row=result_row, column=matrix_idx, value="")
                    
                    cell.border = cell_border
                    cell.alignment = Alignment(vertical="center")
            
            # Instructions
            inst_row = 7 + num_result_rows + 2
            lookup_ws.cell(row=inst_row, column=1, value="Instrucciones:")
            lookup_ws.cell(row=inst_row, column=1).font = Font(bold=True, color="64748B")
            lookup_ws.cell(row=inst_row + 1, column=1, value="1. Selecciona un usuario del menú desplegable en la celda amarilla (B4)")
            lookup_ws.cell(row=inst_row + 1, column=1).font = Font(color="64748B")
            lookup_ws.cell(row=inst_row + 2, column=1, value="2. Los permisos de cada matriz aparecerán automáticamente en columnas separadas")
            lookup_ws.cell(row=inst_row + 2, column=1).font = Font(color="64748B")
            lookup_ws.cell(row=inst_row + 3, column=1, value="3. Cada permiso aparece en una fila diferente para facilitar la lectura")
            lookup_ws.cell(row=inst_row + 3, column=1).font = Font(color="64748B")
            
            # ========== MATRIX SHEETS ==========
            for matrix in matrices:
                # Sanitize sheet name (max 31 chars, no special chars)
                sheet_name = matrix['name'][:31]
                for char in ['\\', '/', '*', '?', ':', '[', ']']:
                    sheet_name = sheet_name.replace(char, '_')
                
                ws = wb.create_sheet(title=sheet_name)
                
                # Header row - Y axis (columns) as column headers
                ws.cell(row=1, column=1, value='')
                ws.cell(row=1, column=1).fill = header_fill
                ws.cell(row=1, column=1).border = cell_border
                
                for col_idx, y_val in enumerate(matrix['yAxis'], start=2):
                    cell = ws.cell(row=1, column=col_idx, value=y_val)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = cell_border
                
                # Data rows - X axis (rows) as row labels
                for row_idx, x_val in enumerate(matrix['xAxis'], start=2):
                    # Row header
                    row_header = ws.cell(row=row_idx, column=1, value=x_val)
                    row_header.font = Font(bold=True)
                    row_header.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
                    row_header.border = cell_border
                    
                    # Data cells
                    x_idx = row_idx - 2
                    if x_idx < len(matrix['data']):
                        for col_idx, val in enumerate(matrix['data'][x_idx], start=2):
                            cell = ws.cell(row=row_idx, column=col_idx, value=val if val == 1 else '')
                            cell.border = cell_border
                            cell.alignment = Alignment(horizontal="center")
                            if val == 1:
                                cell.fill = one_fill
                                cell.font = Font(bold=True, color="16A34A")
                
                # Adjust column widths
                ws.column_dimensions['A'].width = 40
                for col_idx in range(2, len(matrix['yAxis']) + 2):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15
                
                # Freeze panes
                ws.freeze_panes = 'B2'
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'matrices_{timestamp}.xlsx'
            
            self.send_file_download(output.read(), filename)
        except Exception as e:
            self.send_json({'error': str(e)}, 500)


def run_server(port=8080):
    """Start the web server"""
    # Create server with larger request limits
    server = HTTPServer(('127.0.0.1', port), MatrixProcessorHandler)
    server.request_queue_size = 10  # Allow more pending connections
    print(f"\n{'='*50}")
    print(f"  PROCESADOR DE MATRICES")
    print(f"{'='*50}")
    print(f"\n  Servidor ejecutándose en: http://localhost:{port}")
    print(f"\n  Abriendo navegador...")
    print(f"\n  Mantén esta ventana abierta mientras usas la app.")
    print(f"  Presiona Ctrl+C para detener.\n")
    print(f"{'='*50}\n")
    
    # Open browser after a short delay
    threading.Timer(1.0, lambda: webbrowser.open(f'http://localhost:{port}')).start()
    
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServidor detenido.")
        server.shutdown()


if __name__ == '__main__':
    run_server()
